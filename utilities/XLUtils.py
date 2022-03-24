# import lib
import openpyxl

# function return no. of rows
def getRowsCount(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_row

# function return no. of columns
def getColCount(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_column

# function to read data
def ReadData(file, sheet, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.cell(row=row_num, column=col_num).value 

# function to write data
def WriteData(file, sheet, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    sheet.cell(row = row_num, column = col_num).value = data
    workbook.save(file)
